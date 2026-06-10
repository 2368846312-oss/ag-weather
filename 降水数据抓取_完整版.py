#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
全球农产品主产区降水数据抓取脚本（完整版 - 285个坐标）
功能：多线程并行抓取前14天+后14天（共28天，不含当日）日度降水数据
输出：按作物分Sheet的Excel文件，每次运行自动覆盖
需要：pip install openpyxl
"""

import urllib.request
import urllib.parse
import json
import os
import time
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ==================== 配置 ====================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "全球农产品降水数据_最新.xlsx")
THREADS = 16          # 并行线程数，可调整（10~20 均可）
REQUEST_TIMEOUT = 20  # 单个请求超时秒数
MAX_RETRIES = 2       # 失败重试次数

# ==================== 285个主产区坐标 ====================
# 格式: (国家, 作物, 产区, 纬度, 经度, 产量占比)
ALL_REGIONS = [
    # ===== 中国 (36) =====
    ("中国", "大豆", "黑龙江", 46.5, 126.5, "46%"),
    ("中国", "大豆", "吉林", 43.5, 125.0, "8%"),
    ("中国", "大豆", "四川", 30.5, 104.0, "6%"),
    ("中国", "大豆", "安徽", 32.0, 117.0, "5%"),
    ("中国", "大豆", "河南", 34.5, 113.5, "4%"),
    ("中国", "小麦", "河南", 34.5, 113.5, "20%"),
    ("中国", "小麦", "山东", 36.5, 117.0, "19%"),
    ("中国", "小麦", "安徽", 32.0, 117.0, "12%"),
    ("中国", "小麦", "河北", 38.0, 115.0, "11%"),
    ("中国", "小麦", "江苏", 33.0, 119.5, "10%"),
    ("中国", "小麦", "新疆", 44.0, 86.0, "4%"),
    ("中国", "棉花", "新疆", 41.5, 84.0, "87%"),
    ("中国", "玉米", "黑龙江", 46.5, 126.5, "15%"),
    ("中国", "玉米", "吉林", 43.5, 125.0, "12%"),
    ("中国", "玉米", "内蒙古", 43.0, 113.0, "11%"),
    ("中国", "玉米", "山东", 36.5, 117.0, "10%"),
    ("中国", "玉米", "河北", 38.0, 115.0, "8%"),
    ("中国", "玉米", "河南", 34.5, 113.5, "8%"),
    ("中国", "玉米", "辽宁", 41.5, 122.5, "7%"),
    ("中国", "玉米", "山西", 37.5, 112.5, "6%"),
    ("中国", "玉米", "新疆", 41.5, 84.0, "4%"),
    ("中国", "玉米", "四川", 30.5, 104.0, "4%"),
    ("中国", "玉米", "云南", 25.0, 102.0, "4%"),
    ("中国", "菜籽", "四川", 30.5, 104.0, "25%"),
    ("中国", "菜籽", "湖北", 30.5, 113.0, "17%"),
    ("中国", "菜籽", "湖南", 28.0, 112.5, "16%"),
    ("中国", "菜籽", "贵州", 26.5, 106.5, "6%"),
    ("中国", "菜籽", "安徽", 32.0, 117.0, "6%"),
    ("中国", "菜籽", "江西", 28.0, 116.0, "5%"),
    ("中国", "菜籽", "云南", 25.0, 102.0, "4%"),
    ("中国", "菜籽", "重庆", 30.0, 107.0, "4%"),
    ("中国", "菜籽", "江苏", 33.0, 119.5, "4%"),

    # ===== 美国 (34) =====
    ("美国", "大豆", "Illinois", 40.0, -89.0, "16%"),
    ("美国", "大豆", "Iowa", 42.0, -93.5, "14%"),
    ("美国", "大豆", "Minnesota", 44.5, -94.5, "8%"),
    ("美国", "大豆", "Indiana", 40.0, -86.0, "8%"),
    ("美国", "大豆", "Nebraska", 41.0, -98.0, "6%"),
    ("美国", "大豆", "Missouri", 39.0, -92.5, "6%"),
    ("美国", "大豆", "North Dakota", 47.5, -100.0, "5%"),
    ("美国", "大豆", "South Dakota", 44.5, -100.0, "4%"),
    ("美国", "小麦", "Kansas", 38.5, -98.0, "18%"),
    ("美国", "小麦", "North Dakota", 47.5, -100.0, "16%"),
    ("美国", "小麦", "Montana", 47.0, -110.0, "10%"),
    ("美国", "小麦", "Washington", 47.0, -119.0, "9%"),
    ("美国", "小麦", "Oklahoma", 35.5, -97.5, "6%"),
    ("美国", "小麦", "Idaho", 43.0, -114.0, "5%"),
    ("美国", "小麦", "Minnesota", 44.5, -94.5, "5%"),
    ("美国", "棉花", "Texas", 32.0, -100.0, "30%"),
    ("美国", "棉花", "Georgia", 32.5, -83.5, "16%"),
    ("美国", "棉花", "Arkansas", 34.5, -92.0, "9%"),
    ("美国", "棉花", "Mississippi", 33.0, -90.0, "7%"),
    ("美国", "棉花", "Missouri", 39.0, -92.5, "6%"),
    ("美国", "棉花", "North Carolina", 35.5, -78.0, "5%"),
    ("美国", "棉花", "Alabama", 32.5, -86.5, "5%"),
    ("美国", "棉花", "Tennessee", 35.5, -86.5, "4%"),
    ("美国", "玉米", "Iowa", 42.0, -93.5, "18%"),
    ("美国", "玉米", "Illinois", 40.0, -89.0, "16%"),
    ("美国", "玉米", "Minnesota", 44.5, -94.5, "10%"),
    ("美国", "玉米", "Nebraska", 41.0, -98.0, "10%"),
    ("美国", "玉米", "Indiana", 40.0, -86.0, "7%"),
    ("美国", "玉米", "South Dakota", 44.5, -100.0, "5%"),
    ("美国", "玉米", "Ohio", 40.5, -83.0, "5%"),
    ("美国", "玉米", "Wisconsin", 44.0, -90.0, "4%"),
    ("美国", "玉米", "Kansas", 38.5, -98.0, "4%"),
    ("美国", "高粱", "Kansas", 38.5, -98.0, "57%"),
    ("美国", "高粱", "Texas", 35.0, -101.0, "25%"),

    # ===== 巴西 (26) =====
    ("巴西", "大豆", "马托格罗索州", -13.0, -55.0, "28%"),
    ("巴西", "大豆", "巴拉那州", -24.5, -51.5, "13%"),
    ("巴西", "大豆", "戈亚斯州", -16.0, -49.5, "11%"),
    ("巴西", "大豆", "南里奥格兰德州", -29.0, -53.5, "11%"),
    ("巴西", "大豆", "南马托格罗索州", -21.0, -54.5, "9%"),
    ("巴西", "大豆", "米纳斯吉拉斯州", -18.5, -44.5, "6%"),
    ("巴西", "大豆", "巴伊亚州", -12.0, -45.0, "5%"),
    ("巴西", "棉花", "南马托格罗索州", -21.0, -54.5, "73%"),
    ("巴西", "棉花", "巴伊亚州", -12.0, -45.0, "20%"),
    ("巴西", "玉米", "南马托格罗索州", -21.0, -54.5, "37%"),
    ("巴西", "玉米", "巴拉那州", -24.5, -51.5, "13%"),
    ("巴西", "玉米", "戈亚斯州", -16.0, -49.5, "11%"),
    ("巴西", "玉米", "马托格罗索州", -13.0, -55.0, "10%"),
    ("巴西", "玉米", "米纳斯吉拉斯州", -18.5, -44.5, "7%"),
    ("巴西", "玉米", "圣保罗州", -22.0, -48.5, "4%"),
    ("巴西", "甘蔗", "马托格罗索州", -13.0, -55.0, "37%"),
    ("巴西", "甘蔗", "巴拉那州", -24.5, -51.5, "13%"),
    ("巴西", "甘蔗", "戈亚斯州", -16.0, -49.5, "11%"),
    ("巴西", "甘蔗", "南马托格罗索州", -21.0, -54.5, "10%"),
    ("巴西", "甘蔗", "北部/东北部", -8.0, -36.0, "10%"),
    ("巴西", "甘蔗", "米纳斯吉拉斯州", -18.5, -44.5, "7%"),
    ("巴西", "甘蔗", "圣保罗州", -22.0, -48.5, "4%"),

    # ===== 印度甘蔗 (4) =====
    ("印度", "甘蔗", "北方邦", 27.0, 80.5, "48.5%"),
    ("印度", "甘蔗", "马哈拉施特拉邦", 19.5, 76.0, "24%"),
    ("印度", "甘蔗", "卡纳塔克邦", 14.5, 76.0, "10.5%"),
    ("印度", "甘蔗", "泰米尔纳德邦", 11.0, 78.0, "3.3%"),

    # ===== 中国甘蔗 (4) =====
    ("中国", "甘蔗", "广西", 22.5, 107.5, "60%"),
    ("中国", "甘蔗", "云南", 23.5, 99.5, "20%"),
    ("中国", "甘蔗", "广东", 21.4, 110.3, "8%"),
    ("中国", "甘蔗", "海南", 19.5, 109.5, "5%"),

    ("巴西", "高粱", "戈亚斯州", -16.0, -49.5, "40%"),
    ("巴西", "高粱", "米纳斯吉拉斯州", -18.5, -44.5, "25%"),
    ("巴西", "高粱", "圣保罗州", -22.0, -48.5, "10%"),
    ("巴西", "高粱", "马托格罗索州", -13.0, -55.0, "8%"),

    # ===== 阿根廷 (15) =====
    ("阿根廷", "大豆", "Buenos Aires", -36.0, -60.0, "31%"),
    ("阿根廷", "大豆", "Cordoba", -32.5, -63.5, "28%"),
    ("阿根廷", "大豆", "Santa Fe", -31.0, -61.0, "19%"),
    ("阿根廷", "大豆", "Santiago del Estero", -27.5, -63.5, "9%"),
    ("阿根廷", "大麦", "Buenos Aires", -36.0, -60.0, "93%"),
    ("阿根廷", "小麦", "Buenos Aires", -36.0, -60.0, "50%"),
    ("阿根廷", "小麦", "Santa Fe", -31.0, -61.0, "18%"),
    ("阿根廷", "小麦", "Cordoba", -32.5, -63.5, "13%"),
    ("阿根廷", "玉米", "Cordoba", -32.5, -63.5, "35%"),
    ("阿根廷", "玉米", "Buenos Aires", -36.0, -60.0, "27%"),
    ("阿根廷", "玉米", "Santa Fe", -31.0, -61.0, "10%"),
    ("阿根廷", "玉米", "Santiago del Estero", -27.5, -63.5, "9%"),
    ("阿根廷", "葵籽", "Buenos Aires", -36.0, -60.0, "59%"),
    ("阿根廷", "葵籽", "Santa Fe", -31.0, -61.0, "12%"),
    ("阿根廷", "葵籽", "La Pampa", -36.5, -64.5, "12%"),

    # ===== 欧盟 (43) =====
    ("欧盟", "大麦", "Germany", 51.5, 10.5, "33%"),
    ("欧盟", "大麦", "France", 47.0, 2.5, "28%"),
    ("欧盟", "大麦", "Romania", 45.5, 25.0, "5%"),
    ("欧盟", "大麦", "Hungary", 47.0, 19.5, "5%"),
    ("欧盟", "大麦", "Poland", 52.0, 19.0, "4%"),
    ("欧盟", "大麦", "Italy", 43.5, 11.5, "4%"),
    ("欧盟", "小麦", "France", 47.0, 2.5, "27%"),
    ("欧盟", "小麦", "Germany", 51.5, 10.5, "17%"),
    ("欧盟", "小麦", "Poland", 52.0, 19.0, "9%"),
    ("欧盟", "小麦", "Romania", 45.5, 25.0, "7%"),
    ("欧盟", "小麦", "Spain", 40.0, -3.5, "5%"),
    ("欧盟", "小麦", "Bulgaria", 42.5, 25.5, "5%"),
    ("欧盟", "小麦", "United Kingdom", 53.0, -1.5, "5%"),
    ("欧盟", "小麦", "Italy", 43.5, 11.5, "5%"),
    ("欧盟", "小麦", "Czech Republic", 49.5, 15.0, "4%"),
    ("欧盟", "小麦", "Hungary", 47.0, 19.5, "4%"),
    ("欧盟", "玉米", "France", 47.0, 2.5, "21%"),
    ("欧盟", "玉米", "Romania", 45.5, 25.0, "17%"),
    ("欧盟", "玉米", "Poland", 52.0, 19.0, "12%"),
    ("欧盟", "玉米", "Italy", 45.0, 10.0, "9%"),
    ("欧盟", "玉米", "Hungary", 47.0, 19.5, "9%"),
    ("欧盟", "玉米", "Spain", 40.0, -3.5, "7%"),
    ("欧盟", "玉米", "Germany", 51.5, 10.5, "6%"),
    ("欧盟", "玉米", "Bulgaria", 42.5, 25.5, "5%"),
    ("欧盟", "玉米", "Austria", 47.5, 14.0, "4%"),
    ("欧盟", "甜菜", "France", 47.0, 2.5, "30%"),
    ("欧盟", "甜菜", "Germany", 51.5, 10.5, "28%"),
    ("欧盟", "甜菜", "Poland", 52.0, 19.0, "13%"),
    ("欧盟", "甜菜", "Netherlands", 52.0, 5.5, "6%"),
    ("欧盟", "甜菜", "Czech Republic", 49.5, 15.0, "4%"),
    ("欧盟", "菜籽", "Germany", 51.5, 10.5, "21%"),
    ("欧盟", "菜籽", "France", 47.0, 2.5, "21%"),
    ("欧盟", "菜籽", "Poland", 52.0, 19.0, "18%"),
    ("欧盟", "菜籽", "Czech Republic", 49.5, 15.0, "6%"),
    ("欧盟", "菜籽", "Romania", 45.5, 25.0, "6%"),
    ("欧盟", "菜籽", "Lithuania", 55.0, 24.0, "6%"),
    ("欧盟", "菜籽", "Hungary", 47.0, 19.5, "4%"),
    ("欧盟", "菜籽", "Denmark", 55.5, 10.0, "4%"),
    ("欧盟", "葵籽", "Romania", 45.5, 25.0, "29%"),
    ("欧盟", "葵籽", "Bulgaria", 42.5, 25.5, "19%"),
    ("欧盟", "葵籽", "Hungary", 47.0, 19.5, "18%"),
    ("欧盟", "葵籽", "France", 47.0, 2.5, "17%"),
    ("欧盟", "葵籽", "Spain", 40.0, -3.5, "17%"),

    # ===== 俄罗斯 (17) =====
    ("俄罗斯", "大麦", "中央", 54.0, 38.0, "33%"),
    ("俄罗斯", "大麦", "伏尔加", 53.0, 50.0, "29%"),
    ("俄罗斯", "大麦", "南部", 46.0, 41.0, "13%"),
    ("俄罗斯", "大麦", "西伯利亚", 55.0, 80.0, "12%"),
    ("俄罗斯", "大麦", "西北部", 59.0, 31.0, "7%"),
    ("俄罗斯", "大麦", "乌拉尔", 56.5, 60.0, "6%"),
    ("俄罗斯", "大麦", "北高加索", 43.5, 43.0, "5%"),
    ("俄罗斯", "小麦", "南部", 46.0, 41.0, "31%"),
    ("俄罗斯", "小麦", "中央", 54.0, 38.0, "24%"),
    ("俄罗斯", "小麦", "伏尔加", 53.0, 50.0, "18%"),
    ("俄罗斯", "小麦", "西伯利亚", 55.0, 80.0, "13%"),
    ("俄罗斯", "小麦", "北高加索", 43.5, 43.0, "9%"),
    ("俄罗斯", "小麦", "乌拉尔", 56.5, 60.0, "4%"),
    ("俄罗斯", "葵籽", "伏尔加", 53.0, 50.0, "36%"),
    ("俄罗斯", "葵籽", "南部", 46.0, 41.0, "28%"),
    ("俄罗斯", "葵籽", "中央", 54.0, 38.0, "26%"),
    ("俄罗斯", "葵籽", "北高加索", 43.5, 43.0, "4%"),

    # ===== 乌克兰 (30) =====
    ("乌克兰", "玉米", "Zaporizka", 47.5, 35.5, "24%"),
    ("乌克兰", "玉米", "Khersonska", 46.5, 33.5, "21%"),
    ("乌克兰", "玉米", "Kirovohradska", 48.5, 32.0, "20%"),
    ("乌克兰", "玉米", "Dnipropetrovska", 48.5, 35.0, "14%"),
    ("乌克兰", "玉米", "Mykolaivska", 47.0, 31.5, "14%"),
    ("乌克兰", "玉米", "Sumy", 51.0, 34.0, "13%"),
    ("乌克兰", "玉米", "Odeska", 46.5, 30.0, "11%"),
    ("乌克兰", "玉米", "Poltava", 49.5, 34.0, "10%"),
    ("乌克兰", "玉米", "Chernihivska", 51.0, 31.5, "10%"),
    ("乌克兰", "玉米", "Vinnitsa", 49.0, 28.5, "10%"),
    ("乌克兰", "玉米", "Cherkaska", 49.0, 31.5, "8%"),
    ("乌克兰", "玉米", "Donetsk", 48.0, 37.5, "7%"),
    ("乌克兰", "玉米", "Kyivska", 50.0, 30.5, "7%"),
    ("乌克兰", "玉米", "Khmelnitska", 49.5, 27.0, "6%"),
    ("乌克兰", "玉米", "Zhytomyrska", 50.5, 28.5, "5%"),
    ("乌克兰", "葵籽", "Kharkivska", 49.5, 36.5, "12%"),
    ("乌克兰", "葵籽", "Zaporizka", 47.5, 35.5, "10%"),
    ("乌克兰", "葵籽", "Kirovohradska", 48.5, 32.0, "9%"),
    ("乌克兰", "葵籽", "Dnipropetrovska", 48.5, 35.0, "8%"),
    ("乌克兰", "葵籽", "Odeska", 46.5, 30.0, "7%"),
    ("乌克兰", "葵籽", "Poltavska", 49.5, 34.0, "6%"),
    ("乌克兰", "葵籽", "Vinnitska", 49.0, 28.5, "6%"),
    ("乌克兰", "葵籽", "Luhanska", 48.5, 39.0, "5%"),
    ("乌克兰", "葵籽", "Cherkask", 49.0, 31.5, "5%"),
    ("乌克兰", "葵籽", "Sumy", 51.0, 34.0, "5%"),
    ("乌克兰", "葵籽", "Cherkaska", 49.0, 31.5, "4%"),
    ("乌克兰", "葵籽", "Donetsk", 48.0, 37.5, "4%"),
    ("乌克兰", "葵籽", "Khersonska", 46.5, 33.5, "4%"),
    ("乌克兰", "葵籽", "Chivnivska", 48.3, 25.9, "3%"),

    # ===== 加拿大 (7) =====
    ("加拿大", "小麦", "Saskatchewan", 52.0, -106.5, "46%"),
    ("加拿大", "小麦", "Alberta", 53.5, -113.5, "30%"),
    ("加拿大", "小麦", "Manitoba", 50.0, -98.0, "16%"),
    ("加拿大", "小麦", "Ontario", 44.0, -80.0, "8%"),
    ("加拿大", "菜籽", "Saskatchewan", 52.0, -106.5, "54%"),
    ("加拿大", "菜籽", "Alberta", 53.5, -113.5, "29%"),
    ("加拿大", "菜籽", "Manitoba", 50.0, -98.0, "16%"),

    # ===== 印度 (17) =====
    ("印度", "小麦", "北方邦", 27.0, 80.5, "30%"),
    ("印度", "小麦", "中央邦", 23.5, 78.5, "23%"),
    ("印度", "小麦", "旁遮普邦", 30.5, 75.5, "15%"),
    ("印度", "小麦", "哈里亚纳邦", 29.0, 76.0, "10%"),
    ("印度", "小麦", "拉贾斯坦邦", 26.5, 74.0, "10%"),
    ("印度", "棉花", "马哈拉施特拉邦", 19.5, 76.0, "29%"),
    ("印度", "棉花", "古吉拉特邦", 22.5, 72.0, "19%"),
    ("印度", "棉花", "特伦甘纳邦", 17.5, 79.0, "18%"),
    ("印度", "棉花", "拉贾斯坦邦", 26.5, 74.0, "9%"),
    ("印度", "棉花", "卡纳塔克邦", 14.5, 76.0, "8%"),
    ("印度", "棉花", "中央邦", 23.5, 78.5, "7%"),
    ("印度", "棉花", "安得拉邦", 16.0, 80.0, "5%"),
    ("印度", "菜籽", "Rajasthan", 26.5, 74.0, "43%"),
    ("印度", "菜籽", "Haryana", 29.0, 76.0, "12%"),
    ("印度", "菜籽", "Madhya Pradesh", 23.5, 78.5, "10%"),
    ("印度", "菜籽", "Uttar Pradesh", 27.0, 80.5, "10%"),
    ("印度", "菜籽", "West Bengal", 23.0, 88.0, "7%"),

    # ===== 澳洲 (15) =====
    ("澳洲", "大麦", "Western Australia", -32.0, 118.0, "30%"),
    ("澳洲", "大麦", "New South Wales", -33.0, 147.0, "28%"),
    ("澳洲", "大麦", "South Australia", -33.5, 138.0, "20%"),
    ("澳洲", "大麦", "Victoria", -36.0, 143.0, "19%"),
    ("澳洲", "小麦", "New South Wales", -33.0, 147.0, "37%"),
    ("澳洲", "小麦", "Western Australia", -32.0, 118.0, "27%"),
    ("澳洲", "小麦", "South Australia", -33.5, 138.0, "17%"),
    ("澳洲", "小麦", "Victoria", -36.0, 143.0, "13%"),
    ("澳洲", "小麦", "Queensland", -27.0, 148.0, "5%"),
    ("澳洲", "棉花", "New South Wales", -33.0, 147.0, "62%"),
    ("澳洲", "棉花", "Queensland", -27.0, 148.0, "38%"),
    ("澳洲", "菜籽", "Western Australia", -32.0, 118.0, "38%"),
    ("澳洲", "菜籽", "New South Wales", -33.0, 147.0, "33%"),
    ("澳洲", "菜籽", "Victoria", -36.0, 143.0, "19%"),
    ("澳洲", "菜籽", "South Australia", -33.5, 138.0, "10%"),

    # ===== 印度尼西亚 (7) =====
    ("印度尼西亚", "棕榈油", "廖内", 0.5, 102.0, "19%"),
    ("印度尼西亚", "棕榈油", "西加里曼丹", 0.0, 110.0, "14%"),
    ("印度尼西亚", "棕榈油", "中加里曼丹", -2.0, 113.0, "13%"),
    ("印度尼西亚", "棕榈油", "北苏门答腊", 3.0, 98.5, "9%"),
    ("印度尼西亚", "棕榈油", "东加里曼丹", 0.5, 117.0, "9%"),
    ("印度尼西亚", "棕榈油", "南苏门答腊", -3.0, 104.0, "8%"),
    ("印度尼西亚", "棕榈油", "占碑", -1.5, 103.0, "7%"),

    # ===== 马来西亚 (5) =====
    ("马来西亚", "棕榈油", "Sabah", 5.5, 117.0, "24%"),
    ("马来西亚", "棕榈油", "Sarawak", 3.0, 113.5, "21%"),
    ("马来西亚", "棕榈油", "Pahang", 3.5, 103.0, "16%"),
    ("马来西亚", "棕榈油", "Johor", 2.0, 103.5, "16%"),
    ("马来西亚", "棕榈油", "Perak", 4.5, 101.0, "10%"),

    # ===== 泰国 (4) =====
    ("泰国", "棕榈油", "南部", 8.0, 99.5, "90%"),
    ("泰国", "棕榈油", "东北部", 16.0, 103.0, "23%"),
    ("泰国", "棕榈油", "东部", 13.0, 101.5, "21%"),
    ("泰国", "棕榈油", "中部", 14.5, 100.5, "8%"),

    # ===== 巴基斯坦 (4) =====
    ("巴基斯坦", "小麦", "Punjab", 31.0, 73.0, "77%"),
    ("巴基斯坦", "小麦", "Sindh", 26.0, 68.5, "15%"),
    ("巴基斯坦", "小麦", "Khyber Pakhtunkhwa", 34.0, 72.0, "5%"),
    ("巴基斯坦", "小麦", "Balochistan", 29.0, 66.0, "3.5%"),

    # ===== 尼日利亚 (15) =====
    ("尼日利亚", "高粱", "Kano", 12.0, 8.5, "9%"),
    ("尼日利亚", "高粱", "Kebbi", 11.5, 4.0, "9%"),
    ("尼日利亚", "高粱", "Niger", 9.5, 6.0, "8%"),
    ("尼日利亚", "高粱", "Bauchi", 10.5, 10.0, "7%"),
    ("尼日利亚", "高粱", "Kaduna", 10.5, 7.5, "7%"),
    ("尼日利亚", "高粱", "Katsina", 12.5, 7.5, "6%"),
    ("尼日利亚", "高粱", "Zamfara", 12.0, 6.5, "6%"),
    ("尼日利亚", "高粱", "Sokoto", 13.0, 5.0, "6%"),
    ("尼日利亚", "高粱", "Jigawa", 12.0, 9.5, "5%"),
    ("尼日利亚", "高粱", "Borno", 11.5, 13.0, "5%"),
    ("尼日利亚", "高粱", "Gombe", 10.5, 11.0, "5%"),
    ("尼日利亚", "高粱", "Plateau", 9.5, 9.0, "5%"),
    ("尼日利亚", "高粱", "Taraba", 8.0, 11.0, "5%"),
    ("尼日利亚", "高粱", "Yobe", 12.0, 11.5, "4%"),
    ("尼日利亚", "高粱", "Adamawa", 9.0, 12.5, "4%"),

    # ===== 哥伦比亚 (6) =====
    ("哥伦比亚", "咖啡", "内尔哈", 2.5, -75.5, "18%"),
    ("哥伦比亚", "咖啡", "安蒂奥基亚", 6.5, -75.5, "16%"),
    ("哥伦比亚", "咖啡", "托利马", 4.0, -75.0, "12%"),
    ("哥伦比亚", "咖啡", "考卡", 2.5, -76.5, "11%"),
    ("哥伦比亚", "咖啡", "拉希拉", 11.0, -73.0, "5%"),
    ("哥伦比亚", "咖啡", "卡尔达斯", 5.0, -75.5, "5%"),
]

# ==================== 咖啡主产区（14个坐标，温度+降水双维度）====================
COFFEE_REGIONS = [
    # 巴西咖啡 (4)
    ("巴西", "咖啡", "米纳斯吉拉斯州(阿拉比卡)", -19.73, -47.95),
    ("巴西", "咖啡", "米纳斯吉拉斯州(阿拉比卡)", -19.93, -43.93),
    ("巴西", "咖啡", "圣保罗州(阿拉比卡)", -22.81, -47.13),
    ("巴西", "咖啡", "巴伊亚州(罗布斯塔)", -12.15, -45.01),
    # 越南咖啡 (4)
    ("越南", "咖啡", "多乐省(罗布斯塔)", 12.67, 108.03),
    ("越南", "咖啡", "多乐省(罗布斯塔)", 12.98, 108.27),
    ("越南", "咖啡", "多乐省(罗布斯塔)", 13.98, 108.00),
    ("越南", "咖啡", "多乐省(罗布斯塔)", 14.02, 108.71),
    # 哥伦比亚咖啡 (6)
    ("哥伦比亚", "咖啡", "内尔哈", 2.5, -75.5),
    ("哥伦比亚", "咖啡", "安蒂奥基亚", 6.5, -75.5),
    ("哥伦比亚", "咖啡", "托利马", 4.0, -75.0),
    ("哥伦比亚", "咖啡", "考卡", 2.5, -76.5),
    ("哥伦比亚", "咖啡", "拉希拉", 11.0, -73.0),
    ("哥伦比亚", "咖啡", "卡尔达斯", 5.0, -75.5),
]

# ==================== 核心函数 ====================
def fetch_one(country, crop, region, lat, lon):
    """抓取单个坐标的降水数据，失败自动重试"""
    for attempt in range(MAX_RETRIES + 1):
        try:
            params = {
                "latitude": lat,
                "longitude": lon,
                "daily": "precipitation_sum",
                "past_days": 15,
                "forecast_days": 14,
                "timezone": "auto"
            }
            url = "https://api.open-meteo.com/v1/forecast?" + urllib.parse.urlencode(params)
            req = urllib.request.Request(url, headers={"User-Agent": "PrecipScript/1.0"})
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                data = json.loads(resp.read().decode())

            dates = data["daily"]["time"]
            precip = data["daily"]["precipitation_sum"]

            # 定位今日，取前后各14天
            today_str = date.today().isoformat()
            try:
                ti = dates.index(today_str)
                all_dates = dates[max(0, ti-14):ti] + dates[ti+1:ti+15]
                all_precip = precip[max(0, ti-14):ti] + precip[ti+1:ti+15]
            except ValueError:
                all_dates = dates[-14:] + dates[:14]
                all_precip = precip[-14:] + precip[:14]

            return (country, crop, region, lat, lon, all_dates, all_precip, None)

        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(1)
            else:
                return (country, crop, region, lat, lon, None, None, str(e))

def fetch_temp(country, crop, region, lat, lon):
    """抓取单个咖啡产区的日均温度数据，失败自动重试"""
    for attempt in range(MAX_RETRIES + 1):
        try:
            params = {
                "latitude": lat,
                "longitude": lon,
                "daily": "temperature_2m_mean",
                "past_days": 15,
                "forecast_days": 14,
                "timezone": "auto"
            }
            url = "https://api.open-meteo.com/v1/forecast?" + urllib.parse.urlencode(params)
            req = urllib.request.Request(url, headers={"User-Agent": "CoffeeTempScript/1.0"})
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                data = json.loads(resp.read().decode())

            dates = data["daily"]["time"]
            temps = data["daily"]["temperature_2m_mean"]

            today_str = date.today().isoformat()
            try:
                ti = dates.index(today_str)
                all_dates = dates[max(0, ti-14):ti] + dates[ti+1:ti+15]
                all_temps = temps[max(0, ti-14):ti] + temps[ti+1:ti+15]
            except ValueError:
                all_dates = dates[-14:] + dates[:14]
                all_temps = temps[-14:] + temps[:14]

            return (country, crop, region, lat, lon, all_dates, all_temps, None)

        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(1)
            else:
                return (country, crop, region, lat, lon, None, None, str(e))

def generate_excel(results, coffee_temp_results=None):
    """生成最终Excel文件"""
    try:
        import openpyxl
    except ImportError:
        print("正在安装 openpyxl...")
        os.system("pip install openpyxl -q")
        import openpyxl

    from collections import defaultdict

    # 按作物分组
    crop_groups = defaultdict(list)
    for r in results:
        country, crop, region, lat, lon, dates, precip, err = r
        crop_groups[crop].append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 样式
    header_font = openpyxl.styles.Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    header_fill = openpyxl.styles.PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    data_font = openpyxl.styles.Font(name='微软雅黑', size=9)
    border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin', color='D0D0D0'),
        right=openpyxl.styles.Side(style='thin', color='D0D0D0'),
        top=openpyxl.styles.Side(style='thin', color='D0D0D0'),
        bottom=openpyxl.styles.Side(style='thin', color='D0D0D0'))
    center = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    precip_fill = openpyxl.styles.PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    # 获取日期头（取第一个有效数据的日期）
    date_headers = []
    for r in results:
        if r[5] is not None:
            date_headers = r[5]
            break

    for crop in sorted(crop_groups.keys()):
        safe_title = crop[:31]
        ws = wb.create_sheet(title=safe_title)

        headers = ["国家", "产区", "纬度", "经度"] + date_headers + ["历史14天合计", "未来1-7天", "未来8-14天"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        ws.freeze_panes = 'E2'  # 冻结前4列（国家、产区、纬度、经度）
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 9
        ws.column_dimensions['D'].width = 9

        row_idx = 2
        past_sum_col = 5 + len(date_headers)
        future_1_7_col = 6 + len(date_headers)
        future_8_14_col = 7 + len(date_headers)
        for country, crop_name, region, lat, lon, dates, precip, err in sorted(crop_groups[crop], key=lambda x: x[0]):
            ws.cell(row=row_idx, column=1, value=country)
            ws.cell(row=row_idx, column=2, value=region)
            ws.cell(row=row_idx, column=3, value=lat)
            ws.cell(row=row_idx, column=4, value=lon)

            if precip is not None and len(precip) > 0:
                past_sum = round(sum(precip[:14]), 1)
                future_1_7_sum = round(sum(precip[14:21]), 1)
                future_8_14_sum = round(sum(precip[21:28]), 1)

                day_count = min(len(precip), len(date_headers))
                for i in range(day_count):
                    p = precip[i]
                    cell = ws.cell(row=row_idx, column=5+i, value=round(p, 1))
                    if p and p > 0:
                        cell.fill = precip_fill

                ws.cell(row=row_idx, column=past_sum_col, value=past_sum)
                ws.cell(row=row_idx, column=future_1_7_col, value=future_1_7_sum)
                ws.cell(row=row_idx, column=future_8_14_col, value=future_8_14_sum)
            else:
                ws.cell(row=row_idx, column=5, value=f"失败: {err}")

            for col in range(1, len(headers)+1):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = center

            row_idx += 1

    # 失败汇总 Sheet
    # 咖啡温度数据 Sheet（如果提供了）
    if coffee_temp_results and len(coffee_temp_results) > 0:
        ws_temp = wb.create_sheet("咖啡温度")
        temp_headers = ["国家", "产区", "纬度", "经度"] + date_headers + ["过去14天均值", "未来1-7天均值", "未来8-14天均值"]
        ws_temp.append(temp_headers)
        
        for cell in ws_temp[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        
        ws_temp.freeze_panes = 'E2'
        ws_temp.column_dimensions['A'].width = 10
        ws_temp.column_dimensions['B'].width = 25
        ws_temp.column_dimensions['C'].width = 9
        ws_temp.column_dimensions['D'].width = 9
        
        past_mean_col = 5 + len(date_headers)
        future_1_7_mean_col = 6 + len(date_headers)
        future_8_14_mean_col = 7 + len(date_headers)
        
        row_idx = 2
        for country, crop, region, lat, lon, dates, temps, err in coffee_temp_results:
            ws_temp.cell(row=row_idx, column=1, value=country)
            ws_temp.cell(row=row_idx, column=2, value=region)
            ws_temp.cell(row=row_idx, column=3, value=lat)
            ws_temp.cell(row=row_idx, column=4, value=lon)
            
            if temps is not None and len(temps) > 0:
                past_mean = round(sum(temps[:14]) / min(14, len(temps)), 1)
                f1 = temps[14:21]
                f2 = temps[21:28]
                future_1_7_mean = round(sum(f1) / len(f1), 1) if f1 else None
                future_8_14_mean = round(sum(f2) / len(f2), 1) if f2 else None
                
                day_count = min(len(temps), len(date_headers))
                for i in range(day_count):
                    t = round(temps[i], 1)
                    ws_temp.cell(row=row_idx, column=5+i, value=t)
                
                ws_temp.cell(row=row_idx, column=past_mean_col, value=past_mean if past_mean is not None else "")
                ws_temp.cell(row=row_idx, column=future_1_7_mean_col, value=future_1_7_mean if future_1_7_mean is not None else "")
                ws_temp.cell(row=row_idx, column=future_8_14_mean_col, value=future_8_14_mean if future_8_14_mean is not None else "")
            else:
                ws_temp.cell(row=row_idx, column=5, value=f"失败: {err}")
            
            for col in range(1, len(temp_headers)+1):
                cell = ws_temp.cell(row=row_idx, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = center
            
            row_idx += 1

    failed = [r for r in results if r[5] is None]
    if failed:
        ws_fail = wb.create_sheet("抓取失败")
        ws_fail.append(["国家", "作物", "产区", "错误信息"])
        for row_idx, r in enumerate(failed, 2):
            ws_fail.cell(row=row_idx, column=1, value=r[0])
            ws_fail.cell(row=row_idx, column=2, value=r[1])
            ws_fail.cell(row=row_idx, column=3, value=r[2])
            ws_fail.cell(row=row_idx, column=4, value=r[7])

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE

# ==================== 主流程 ====================
def main():
    print("=" * 60)
    print(f"  全球农产品主产区降水数据抓取")
    print(f"  运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  坐标总数: {len(ALL_REGIONS)}")
    print(f"  并行线程: {THREADS}")
    print(f"  输出文件: {OUTPUT_FILE}")
    print("=" * 60)

    results = []
    ok = fail = 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=THREADS) as executor:
        futures = {}
        for item in ALL_REGIONS:
            country, crop, region, lat, lon, pct = item
            future = executor.submit(fetch_one, country, crop, region, lat, lon)
            futures[future] = item

        for i, future in enumerate(as_completed(futures), 1):
            item = futures[future]
            country, crop, region, lat, lon, pct = item
            try:
                result = future.result()
                results.append(result)
                if result[5] is not None:
                    ok += 1
                    elapsed = time.time() - start_time
                    print(f"[{i}/{len(ALL_REGIONS)}] OK  {country}-{crop}-{region}  ({len(result[5])}天)  |  耗时 {elapsed:.0f}s")
                else:
                    fail += 1
                    print(f"[{i}/{len(ALL_REGIONS)}] FAIL {country}-{crop}-{region}  |  {result[7]}")
            except Exception as e:
                fail += 1
                results.append((country, crop, region, lat, lon, None, None, str(e)))
                print(f"[{i}/{len(ALL_REGIONS)}] FAIL {country}-{crop}-{region}  |  {e}")

    elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  降水抓取完成: {ok}成功 / {fail}失败 / {len(ALL_REGIONS)}总数")
    print(f"  耗时: {elapsed:.0f} 秒")
    print(f"{'=' * 60}")

    # ---- 咖啡产区温度抓取 ----
    coffee_temp_results = []
    if COFFEE_REGIONS:
        ok2 = fail2 = 0
        print(f"\n=== 咖啡产区温度抓取（{len(COFFEE_REGIONS)}个坐标）===")
        t_start = time.time()
        with ThreadPoolExecutor(max_workers=THREADS) as executor:
            tfutures = {}
            for item in COFFEE_REGIONS:
                country, crop, region, lat, lon = item
                future = executor.submit(fetch_temp, country, crop, region, lat, lon)
                tfutures[future] = item

            for i, future in enumerate(as_completed(tfutures), 1):
                item = tfutures[future]
                country, crop, region, lat, lon = item
                try:
                    result = future.result()
                    coffee_temp_results.append(result)
                    if result[5] is not None:
                        ok2 += 1
                        print(f"[{i}/{len(COFFEE_REGIONS)}] TEMP OK  {country}-{region}  ({len(result[5])}天)")
                    else:
                        fail2 += 1
                        print(f"[{i}/{len(COFFEE_REGIONS)}] TEMP FAIL {country}-{region}  |  {result[7]}")
                except Exception as e:
                    fail2 += 1
                    coffee_temp_results.append((country, crop, region, lat, lon, None, None, str(e)))
                    print(f"[{i}/{len(COFFEE_REGIONS)}] TEMP FAIL {country}-{region}  |  {e}")
        t_elapsed = time.time() - t_start
        print(f"  温度抓取完成: {ok2}成功 / {fail2}失败  |  耗时 {t_elapsed:.0f}s")
    else:
        coffee_temp_results = None

    if ok > 0:
        out_path = generate_excel(results, coffee_temp_results)
        print(f"\n  Excel已保存: {out_path}")
    else:
        print("  未获取到任何降水数据，跳过Excel生成")

    print("=" * 60)

if __name__ == "__main__":
    main()
