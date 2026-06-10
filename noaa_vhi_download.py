#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
NOAA STAR VHI 批量并行下载脚本 (v3)
- 并行下载，6线程
- 剔除咖啡、巴基斯坦、泰国
- 巴西/阿根廷省份名直接ID匹配
- 俄罗斯宏观区域 → 多联邦主体聚合（取均值）
- 仅使用标准库
"""

import urllib.request
import urllib.error
import csv
import os
import sys
import json
import re
import time
import threading
from collections import defaultdict
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================================================
# 配置区
# ============================================================
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "VHI_Data")
API_BASE = "https://www.star.nesdis.noaa.gov/smcd/emb/vci/VH/get_TS_admin.php"
CACHE_FILE = os.path.join(OUTPUT_DIR, "province_cache.json")
MAX_WORKERS = 4
REQUEST_TIMEOUT = 30
REQUEST_DELAY = 0.5  # 请求间隔（秒），避免触发限流
print_lock = threading.Lock()

# ============================================================
# 国家代码
# ============================================================
COUNTRY_CODE = {
    "中国": "CHN", "美国": "USA", "巴西": "BRA", "阿根廷": "ARG",
    "俄罗斯": "RUS", "加拿大": "CAN", "印度": "IND",
    "澳洲": "AUS", "乌克兰": "UKR", "尼日利亚": "NGA",
    "印度尼西亚": "IDN", "马来西亚": "MYS", "哥伦比亚": "COL", "越南": "VNM",
    # === 欧盟主要农业国 ===
    "法国": "FRA", "德国": "DEU", "波兰": "POL", "罗马尼亚": "ROU",
    "西班牙": "ESP", "意大利": "ITA", "保加利亚": "BGR", "匈牙利": "HUN",
    "丹麦": "DNK", "捷克": "CZE", "英国": "GBR", "奥地利": "AUT", "立陶宛": "LTU",
}

# ============================================================
# 作物代码（咖啡已剔除）
# ============================================================
CROP_TAG = {
    "大豆": "SOYB", "小麦": "WHEA", "玉米": "MAIZ",
    "棉花": "COTT", "菜籽": "RAPE", "高粱": "SORG",
    "大麦": "BARL", "葵籽": "SUNF", "甘蔗": "SUGC",
    "棕榈油": "OILP", "甜菜": "SUGB",
}

# ============================================================
# 美国省份ID
# ============================================================
USA_PROVINCES = {
    "Alabama": 1, "Alaska": 2, "Arizona": 3, "Arkansas": 4,
    "California": 5, "Colorado": 6, "Connecticut": 7, "Delaware": 8,
    "District of Columbia": 9, "Florida": 10, "Georgia": 11, "Hawaii": 12,
    "Idaho": 13, "Illinois": 14, "Indiana": 15, "Iowa": 16,
    "Kansas": 17, "Kentucky": 18, "Louisiana": 19, "Maine": 20,
    "Maryland": 21, "Massachusetts": 22, "Michigan": 23, "Minnesota": 24,
    "Mississippi": 25, "Missouri": 26, "Montana": 27, "Nebraska": 28,
    "Nevada": 29, "New Hampshire": 30, "New Jersey": 31, "New Mexico": 32,
    "New York": 33, "North Carolina": 34, "North Dakota": 35, "Ohio": 36,
    "Oklahoma": 37, "Oregon": 38, "Pennsylvania": 39, "Rhode Island": 40,
    "South Carolina": 41, "South Dakota": 42, "Tennessee": 43, "Texas": 44,
    "Utah": 45, "Vermont": 46, "Virginia": 47, "Washington": 48,
    "West Virginia": 49, "Wisconsin": 50, "Wyoming": 51,
}

# ============================================================
# 巴西/阿根廷省份ID直接映射（解决缓存编码问题）
# ============================================================
DIRECT_PROVINCE_IDS = {
    ("BRA", "巴拉那州"): 16,      # Paraná
    ("BRA", "戈亚斯州"): 9,       # Goiás
    ("BRA", "圣保罗州"): 25,      # São Paulo
    ("ARG", "Cordoba"): 6,        # Córdoba
}

# ============================================================
# 俄罗斯宏观区域 → NOAA 联邦主体ID列表
#（联邦主体名称含特殊字符，直接用ID避免匹配问题）
# ============================================================
RUS_REGION_MAP = {
    # 南部联邦区（小麦/大麦/葵籽）：Krasnodar(39), Astrakhan'(12)
    "南部": [39, 12],
    # 中央联邦区（小麦/大麦/葵籽）：Belgorod(14), Bryansk(15), Ivanovo(28),
    #   Kaluga(32), Kostroma(38), Kursk(42), Lipetsk(44), Moskva(48), Orel(56)
    "中央": [14, 15, 28, 32, 38, 42, 44, 48, 56],
    # 伏尔加联邦区：Bashkortostan(13), Chuvash(21), Kirov(34),
    #   Mariy-El(46), Mordovia(47), Nizhegorod(51), Orenburg(57), Penza(58)
    "伏尔加": [13, 21, 34, 46, 47, 51, 57, 58],
    # 西伯利亚联邦区：Altay(9), Irkutsk(27), Kemerovo(3),
    #   Khakass(5), Krasnoyarsk(40), Novosibirsk(54), Omsk(55)
    "西伯利亚": [3, 5, 9, 27, 40, 54, 55],
    # 北高加索联邦区：Chechnya(17), Dagestan(23), Ingush(26),
    #   Kabardin-Balkar(29), Karachay-Cherkess(1), North Ossetia(52)
    "北高加索": [1, 17, 23, 26, 29, 52],
    # 乌拉尔联邦区：Chelyabinsk(18), Khanty-Mansiy(6), Kurgan(41), Perm'(59)
    "乌拉尔": [6, 18, 41, 59],
    # 西北部联邦区：Arkhangel'sk(11), City of St. Petersburg(22),
    #   Kaliningrad(30), Karelia(2), Komi(35), Leningrad(43),
    #   Murmansk(49), Nenets(50), Novgorod(53)
    "西北部": [2, 11, 22, 30, 35, 43, 49, 50, 53],
}

# ============================================================
# 欧盟国家级查询（走国家均值API，不查省份）
# ============================================================
EU_COUNTRIES = {"法国", "德国", "波兰", "罗马尼亚", "西班牙", "意大利", "保加利亚", "匈牙利", "丹麦", "捷克", "英国", "奥地利", "立陶宛"}

# 欧盟国家中→英映射（省份列输出用）
EU_CN_TO_EN = {
    "法国": "France", "德国": "Germany", "波兰": "Poland",
    "罗马尼亚": "Romania", "西班牙": "Spain", "意大利": "Italy",
    "保加利亚": "Bulgaria", "匈牙利": "Hungary", "丹麦": "Denmark", "捷克": "Czech",
    "英国": "United Kingdom", "奥地利": "Austria", "立陶宛": "Lithuania",
}

# 所有国家中→英映射（国家列输出用）
CN_TO_EN_COUNTRY = {
    "中国": "China", "美国": "USA", "巴西": "Brazil", "阿根廷": "Argentina",
    "俄罗斯": "Russia", "加拿大": "Canada", "印度": "India",
    "澳洲": "Australia", "乌克兰": "Ukraine", "尼日利亚": "Nigeria",
    "印度尼西亚": "Indonesia", "马来西亚": "Malaysia", "哥伦比亚": "Colombia", "越南": "Vietnam",
    **EU_CN_TO_EN, "欧盟": "EU",
}

# ============================================================
# 中文省份名 → 英文名（用于缓存查找）
# ============================================================
CN_TO_EN_PROVINCE = {
    "黑龙江": "Heilongjiang", "吉林": "Jilin", "四川": "Sichuan", "安徽": "Anhui",
    "河南": "Henan", "山东": "Shandong", "河北": "Hebei", "江苏": "Jiangsu",
    "新疆": "Xinjiang Uygur", "内蒙古": "Nei Mongol", "辽宁": "Liaoning",
    "山西": "Shanxi", "云南": "Yunnan", "湖北": "Hubei", "湖南": "Hunan",
    "贵州": "Guizhou", "江西": "Jiangxi", "重庆": "Chongqing", "广西": "Guangxi",
    "广东": "Guangdong",
    "马托格罗索州": "Mato Grosso",
    "南里奥格兰德州": "Rio Grande do Sul",
    "南马托格罗索州": "Mato Grosso do Sul", "米纳斯吉拉斯州": "Minas Gerais",
    "巴伊亚州": "Bahia",
    "北方邦": "Uttar Pradesh", "中央邦": "Madhya Pradesh", "旁遮普邦": "Punjab",
    "哈里亚纳邦": "Haryana", "拉贾斯坦邦": "Rajasthan", "马哈拉施特拉邦": "Maharashtra",
    "古吉拉特邦": "Gujarat", "特伦甘纳邦": "Telangana", "卡纳塔克邦": "Karnataka",
    "安得拉邦": "Andhra Pradesh", "泰米尔纳德邦": "Tamil Nadu",
    "廖内": "Riau", "西加里曼丹": "Kalimantan Barat", "中加里曼丹": "Kalimantan Tengah",
    "北苏门答腊": "Sumatera Utara", "东加里曼丹": "Kalimantan Timur",
    "南苏门答腊": "Sumatera Selatan", "占碑": "Jambi",
    "Zaporizka": "Zaporizhzhya", "Dnipropetrovska": "Dnipropetrovs'k",
    "Mykolaivska": "Mykolayiv", "Odeska": "Odessa", "Vinnitsa": "Vinnytsya",
    "Cherkaska": "Cherkasy", "Kyivska": "Kiev", "Khmelnitska": "Khmel'nyts'kyy",
    "Luhanska": "Luhans'k", "Poltavska": "Poltava",
    "Chernihivska": "Chernihiv", "Kharkivska": "Kharkiv",
    "Kirovohradska": "Kirovohrad", "Khersonska": "Kherson",
    "Zhytomyrska": "Zhytomyr", "Kirovogradska": "Kirovohrad",
    "Chivnivska": "Chernivtsi",
    "内尔哈": "Huila", "安蒂奥基亚": "Antioquia", "托利马": "Tolima",
    "考卡": "Cauca", "拉希拉": "La Guajira", "卡尔达斯省": "Caldas",
    "多乐省": "Dak Lak", "林同省": "Lam Dong",
}


# ============================================================
# 工具函数
# ============================================================

def load_province_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def find_province_id(cache, country_code, province_name):
    """根据省份名查找NOAA ID。先查直接映射表，再查缓存"""
    # 先查直接映射（解决缓存编码问题）
    if (country_code, province_name) in DIRECT_PROVINCE_IDS:
        return DIRECT_PROVINCE_IDS[(country_code, province_name)]

    provinces = cache.get(country_code, {})
    if not provinces:
        return None

    en_name = CN_TO_EN_PROVINCE.get(province_name, province_name)
    names_to_try = [en_name, province_name]

    for try_name in names_to_try:
        if try_name in provinces:
            return provinces[try_name]
        # 模糊匹配
        pn_lower = try_name.lower().replace(' ', '').replace("'", "")
        for name, pid in provinces.items():
            name_lower = name.lower().replace(' ', '').replace("'", "")
            if pn_lower == name_lower:
                return pid
            if pn_lower in name_lower or name_lower in pn_lower:
                return pid
    return None


def download_vhi(country_code, province_id, crop_tag, year1, year2):
    """下载VHI数据。province_id为None时查询国家级均值"""
    time.sleep(REQUEST_DELAY)
    if province_id is None:
        url = (f"{API_BASE}?country={country_code}"
               f"&yearlyTag=Weekly&type=Mean&TagCropland={crop_tag}"
               f"&year1={year1}&year2={year2}")
    else:
        url = (f"{API_BASE}?country={country_code}&provinceID={province_id}"
               f"&yearlyTag=Weekly&type=Mean&TagCropland={crop_tag}"
               f"&year1={year1}&year2={year2}")
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
            return resp.read().decode('utf-8', errors='replace')
    except urllib.error.HTTPError as e:
        return None if e.code == 404 else None
    except Exception:
        return None


def parse_vhi_data(raw_text):
    results = []
    lines = raw_text.strip().split('\n')
    data_start = False
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith('Mean data') or line.startswith('year,week') or line.startswith('for'):
            if 'year,week' in line:
                data_start = True
            continue
        if not data_start:
            continue
        if line.startswith('20'):
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 7:
                try:
                    year, week = int(parts[0]), int(parts[1])
                    vhi = float(parts[6])
                    if vhi > -0.5:
                        results.append((
                            year, week,
                            float(parts[2]), float(parts[3]),
                            float(parts[4]), float(parts[5]), vhi
                        ))
                except (ValueError, IndexError):
                    continue
    return results


def get_last_6_weeks(data):
    if not data:
        return []
    sorted_data = sorted(data, key=lambda x: (x[0], x[1]), reverse=True)
    return sorted(sorted_data[:6], key=lambda x: (x[0], x[1]))


def aggregate_region_data(all_province_data):
    """
    聚合多个省份的数据（按周取均值）。
    all_province_data: [province_data, ...]
      每个 province_data = [(year, week, smn, smt, vci, tci, vhi), ...]
    返回: [(year, week, avg_smn, avg_smt, avg_vci, avg_tci, avg_vhi), ...]
    """
    week_vals = defaultdict(list)
    for pdata in all_province_data:
        for entry in pdata:
            week_vals[(entry[0], entry[1])].append(entry)

    aggregated = []
    for (year, week), entries in sorted(week_vals.items()):
        n = len(entries)
        avg = (
            year, week,
            round(sum(e[2] for e in entries) / n, 4),  # SMN
            round(sum(e[3] for e in entries) / n, 2),  # SMT
            round(sum(e[4] for e in entries) / n, 2),  # VCI
            round(sum(e[5] for e in entries) / n, 2),  # TCI
            round(sum(e[6] for e in entries) / n, 2),  # VHI
        )
        aggregated.append(avg)
    return aggregated


# ============================================================
# 主产区数据（剔除咖啡、巴基斯坦、泰国）
# ============================================================
def parse_excel_manually():
    return [
        # === 大豆 ===
        ("大豆", "中国", ["黑龙江", "吉林", "四川", "安徽", "河南"]),
        ("大豆", "美国", ["Illinois", "Iowa", "Minnesota", "Indiana", "Nebraska", "Missouri", "North Dakota", "South Dakota"]),
        ("大豆", "巴西", ["马托格罗索州", "巴拉那州", "戈亚斯州", "南里奥格兰德州", "南马托格罗索州", "米纳斯吉拉斯州", "巴伊亚州"]),
        ("大豆", "阿根廷", ["Buenos Aires", "Cordoba", "Santa Fe", "Santiago del Estero"]),

        # === 小麦 ===
        ("小麦", "中国", ["河南", "山东", "安徽", "河北", "江苏", "新疆"]),
        ("小麦", "美国", ["Kansas", "North Dakota", "Montana", "Washington", "Oklahoma", "Idaho", "Minnesota"]),
        ("小麦", "阿根廷", ["Buenos Aires", "Santa Fe", "Cordoba"]),
        ("小麦", "俄罗斯", ["南部", "中央", "伏尔加", "西伯利亚", "北高加索", "乌拉尔"]),
        ("小麦", "加拿大", ["Saskatchewan", "Alberta", "Manitoba", "Ontario"]),
        ("小麦", "印度", ["北方邦", "中央邦", "旁遮普邦", "哈里亚纳邦", "拉贾斯坦邦"]),
        ("小麦", "澳洲", ["New South Wales", "Western Australia", "South Australia", "Victoria", "Queensland"]),

        # === 玉米 ===
        ("玉米", "中国", ["黑龙江", "吉林", "内蒙古", "山东", "河北", "河南", "辽宁", "山西", "新疆", "四川", "云南"]),
        ("玉米", "美国", ["Iowa", "Illinois", "Minnesota", "Nebraska", "Indiana", "South Dakota", "Ohio", "Wisconsin", "Kansas"]),
        ("玉米", "巴西", ["南马托格罗索州", "巴拉那州", "戈亚斯州", "马托格罗索州", "米纳斯吉拉斯州", "圣保罗州"]),
        ("玉米", "阿根廷", ["Cordoba", "Buenos Aires", "Santa Fe", "Santiago del Estero"]),
        ("玉米", "乌克兰", ["Zaporizka", "Khersonska", "Kirovohradska", "Dnipropetrovska", "Mykolaivska",
                           "Sumy", "Odeska", "Poltava", "Chernihivska", "Vinnitsa",
                           "Cherkaska", "Donetsk", "Kyivska", "Khmelnitska", "Zhytomyrska"]),

        # === 棉花 ===
        ("棉花", "中国", ["新疆"]),
        ("棉花", "美国", ["Texas", "Georgia", "Arkansas", "Mississippi", "Missouri", "North Carolina", "Alabama", "Tennessee"]),
        ("棉花", "巴西", ["南马托格罗索州", "巴伊亚州"]),
        ("棉花", "印度", ["马哈拉施特拉邦", "古吉拉特邦", "特伦甘纳邦", "拉贾斯坦邦", "卡纳塔克邦", "中央邦", "安得拉邦"]),
        ("棉花", "澳洲", ["New South Wales", "Queensland"]),

        # === 菜籽 ===
        ("菜籽", "中国", ["四川", "湖北", "湖南", "贵州", "安徽", "江西", "云南", "重庆", "江苏"]),
        ("菜籽", "加拿大", ["Saskatchewan", "Alberta", "Manitoba"]),
        ("菜籽", "印度", ["Rajasthan", "Haryana", "Madhya Pradesh", "Uttar Pradesh", "West Bengal"]),
        ("菜籽", "澳洲", ["Western Australia", "New South Wales", "Victoria", "South Australia"]),

        # === 高粱 ===
        ("高粱", "美国", ["Kansas", "Texas"]),
        ("高粱", "巴西", ["戈亚斯州", "米纳斯吉拉斯州", "圣保罗州", "马托格罗索州"]),
        ("高粱", "尼日利亚", ["Kano", "Kebbi", "Niger", "Bauchi", "Kaduna", "Katsina",
                            "Zamfara", "Sokoto", "Jigawa", "Borno", "Gombe",
                            "Plateau", "Taraba", "Yobe", "Adamawa"]),

        # === 大麦 ===
        ("大麦", "阿根廷", ["Buenos Aires"]),
        ("大麦", "俄罗斯", ["中央", "伏尔加", "南部", "西伯利亚", "西北部", "乌拉尔", "北高加索"]),
        ("大麦", "澳洲", ["Western Australia", "New South Wales", "South Australia", "Victoria"]),

        # === 葵籽 ===
        ("葵籽", "阿根廷", ["Buenos Aires", "Santa Fe", "La Pampa"]),
        ("葵籽", "俄罗斯", ["伏尔加", "南部", "中央", "北高加索"]),
        ("葵籽", "乌克兰", ["Kharkivska", "Zaporizka", "Kirovohradska", "Dnipropetrovska",
                           "Odeska", "Poltavska", "Vinnitsa", "Luhanska",
                           "Cherkask", "Kirovogradska", "Sumy", "Cherkaska",
                           "Donetsk", "Khersonska", "Chivnivska"]),

        # === 甘蔗 ===
        ("甘蔗", "巴西", ["马托格罗索州", "巴拉那州", "戈亚斯州", "南马托格罗索州", "米纳斯吉拉斯州", "圣保罗州"]),
        ("甘蔗", "印度", ["马哈拉施特拉邦", "北方邦", "卡纳塔克邦", "泰米尔纳德邦", "古吉拉特邦", "安得拉邦"]),
        ("甘蔗", "中国", ["广东", "云南", "广西", "贵州"]),

        # === 棕榈油 ===
        ("棕榈油", "印度尼西亚", ["廖内", "西加里曼丹", "中加里曼丹", "北苏门答腊", "东加里曼丹", "南苏门答腊", "占碑"]),
        ("棕榈油", "马来西亚", ["Sabah", "Sarawak", "Pahang", "Johor", "Perak"]),

        # ========== 欧盟主要农业国（国家级查询） ==========
        ("小麦", "法国", ["全国"]), ("小麦", "德国", ["全国"]), ("小麦", "波兰", ["全国"]),
        ("小麦", "罗马尼亚", ["全国"]), ("小麦", "西班牙", ["全国"]), ("小麦", "意大利", ["全国"]),
        ("小麦", "保加利亚", ["全国"]), ("小麦", "匈牙利", ["全国"]),
        ("小麦", "英国", ["全国"]), ("小麦", "捷克", ["全国"]),
        ("玉米", "法国", ["全国"]), ("玉米", "罗马尼亚", ["全国"]), ("玉米", "匈牙利", ["全国"]),
        ("玉米", "德国", ["全国"]), ("玉米", "意大利", ["全国"]), ("玉米", "波兰", ["全国"]),
        ("玉米", "西班牙", ["全国"]), ("玉米", "保加利亚", ["全国"]), ("玉米", "奥地利", ["全国"]),
        ("菜籽", "法国", ["全国"]), ("菜籽", "德国", ["全国"]), ("菜籽", "波兰", ["全国"]),
        ("菜籽", "罗马尼亚", ["全国"]), ("菜籽", "捷克", ["全国"]),
        ("菜籽", "匈牙利", ["全国"]), ("菜籽", "立陶宛", ["全国"]), ("菜籽", "丹麦", ["全国"]),
        ("大麦", "法国", ["全国"]), ("大麦", "德国", ["全国"]), ("大麦", "西班牙", ["全国"]),
        ("大麦", "丹麦", ["全国"]), ("大麦", "罗马尼亚", ["全国"]), ("大麦", "匈牙利", ["全国"]),
        ("大麦", "波兰", ["全国"]), ("大麦", "意大利", ["全国"]),
        ("葵籽", "法国", ["全国"]), ("葵籽", "罗马尼亚", ["全国"]), ("葵籽", "保加利亚", ["全国"]),
        ("葵籽", "匈牙利", ["全国"]), ("葵籽", "西班牙", ["全国"]),
    ]


# ============================================================
# 下载任务
# ============================================================

def download_single(crop_name, country_name, province, country_code, province_id, crop_tag, current_year):
    """下载单个省份-作物组合的数据"""
    try:
        raw = download_vhi(country_code, province_id, crop_tag, current_year - 1, current_year)
        if raw is None:
            return (False, f"[空数据] {country_name}/{province} ({crop_name})", None)

        data = parse_vhi_data(raw)
        if not data:
            return (False, f"[无有效数据] {country_name}/{province} ({crop_name})", None)

        last6 = get_last_6_weeks(data)
        results = []
        for entry in last6:
            results.append({
                "作物": crop_name, "国家": country_name, "省份": province,
                "年份": entry[0], "周": entry[1],
                "VHI": entry[6],  # 只保留 VHI 值
            })
        latest = last6[-1]
        return (True, f"[成功] {country_name}/{province} ({crop_name}): "
                     f"{len(last6)}周, 最新 {latest[0]} W{latest[1]} VHI={latest[6]:.2f}", results)
    except Exception as e:
        return (False, f"[异常] {country_name}/{province} ({crop_name}): {e}", None)


def download_rus_region(crop_name, region_name, province_ids, crop_tag, current_year):
    """下载俄罗斯宏观区域（聚合多个联邦主体）"""
    try:
        all_data = []
        success_ids = []
        for pid in province_ids:
            raw = download_vhi("RUS", pid, crop_tag, current_year - 1, current_year)
            if raw:
                data = parse_vhi_data(raw)
                if data:
                    all_data.append(data)
                    success_ids.append(pid)

        if not all_data:
            return (False, f"[空数据] 俄罗斯/{region_name} ({crop_name}): 所有联邦主体均无数据", None)

        # 聚合
        aggregated = aggregate_region_data(all_data)
        last6 = get_last_6_weeks(aggregated)

        results = []
        for entry in last6:
            results.append({
                "作物": crop_name, "国家": "俄罗斯", "省份": region_name,
                "年份": entry[0], "周": entry[1],
                "VHI": entry[6],  # 只保留 VHI 值
            })
        latest = last6[-1]
        return (True, f"[成功] 俄罗斯/{region_name} ({crop_name}): "
                     f"{len(last6)}周 ({len(success_ids)}/{len(province_ids)}主体), "
                     f"最新 {latest[0]} W{latest[1]} VHI={latest[6]:.2f}", results)
    except Exception as e:
        return (False, f"[异常] 俄罗斯/{region_name} ({crop_name}): {e}", None)


def download_country(crop_name, country_name, country_code, crop_tag, current_year):
    """下载国家级VHI均值（用于无省份ID的国家，如欧盟各国）"""
    try:
        raw = download_vhi(country_code, None, crop_tag, current_year - 1, current_year)
        if raw is None:
            return (False, f"[空数据] {country_name} ({crop_name})", None)

        data = parse_vhi_data(raw)
        if not data:
            return (False, f"[无有效数据] {country_name} ({crop_name})", None)

        last6 = get_last_6_weeks(data)
        results = []
        for entry in last6:
            results.append({
                "作物": crop_name, "国家": "欧盟", "省份": EU_CN_TO_EN.get(country_name, country_name),
                "年份": entry[0], "周": entry[1],
                "VHI": entry[6],
            })
        latest = last6[-1]
        return (True, f"[成功] {country_name} ({crop_name}): "
                     f"{len(last6)}周, 最新 {latest[0]} W{latest[1]} VHI={latest[6]:.2f}", results)
    except Exception as e:
        return (False, f"[异常] {country_name} ({crop_name}): {e}", None)


def download_one_task(task):
    """线程池任务入口"""
    task_type = task[0]
    if task_type == "normal":
        _, crop_name, country_name, province, country_code, province_id, crop_tag, current_year = task
        return download_single(crop_name, country_name, province, country_code, province_id, crop_tag, current_year)
    elif task_type == "rus_region":
        _, crop_name, region_name, province_ids, crop_tag, current_year = task
        return download_rus_region(crop_name, region_name, province_ids, crop_tag, current_year)
    elif task_type == "country":
        _, crop_name, country_name, country_code, crop_tag, current_year = task
        return download_country(crop_name, country_name, country_code, crop_tag, current_year)


# ============================================================
# 主流程
# ============================================================

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 加载缓存
    cache = load_province_cache()
    total_cached = sum(len(v) for v in cache.values())
    print(f"[缓存] {len(cache)}国 {total_cached} 省份已加载")
    sys.stdout.flush()

    current_year = datetime.now().year
    summary_path = os.path.join(OUTPUT_DIR, "VHI_最近6周汇总.xlsx")

    crop_areas = parse_excel_manually()

    # 构建任务
    normal_tasks = []
    rus_tasks = []
    country_tasks = []
    skip_records = []

    for crop_name, country_name, provinces in crop_areas:
        country_code = COUNTRY_CODE.get(country_name)
        crop_tag = CROP_TAG.get(crop_name)
        if not country_code or not crop_tag:
            continue

        for province in provinces:
            # 俄罗斯走区域聚合路径
            if country_code == "RUS":
                if province in RUS_REGION_MAP:
                    rus_tasks.append(("rus_region", crop_name, province, RUS_REGION_MAP[province], crop_tag, current_year))
                else:
                    skip_records.append(f"[跳过] 俄罗斯未知区域: {province} ({crop_name})")
                continue

            # 欧盟国家级查询（不查省份）
            if country_name in EU_COUNTRIES:
                country_tasks.append(("country", crop_name, country_name, country_code, crop_tag, current_year))
                continue

            # 美国
            if country_code == "USA":
                province_id = USA_PROVINCES.get(province)
            else:
                province_id = find_province_id(cache, country_code, province)

            if province_id is None:
                skip_records.append(f"[跳过] 未找到ID: {country_name}/{province} ({crop_name})")
                continue

            normal_tasks.append(("normal", crop_name, country_name, province, country_code, province_id, crop_tag, current_year))

    total_tasks = len(normal_tasks) + len(rus_tasks) + len(country_tasks)
    print(f"普通任务: {len(normal_tasks)}, 俄罗斯聚合任务: {len(rus_tasks)}, 国家级任务: {len(country_tasks)}, 跳过: {len(skip_records)}")
    for s in skip_records:
        print(f"  {s}")
    sys.stdout.flush()

    all_results = []
    success_count = 0
    fail_count = 0
    completed = 0

    all_tasks = normal_tasks + rus_tasks + country_tasks

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(download_one_task, t): t for t in all_tasks}

        for future in as_completed(future_map):
            completed += 1
            try:
                success, msg, results = future.result()
            except Exception as e:
                with print_lock:
                    print(f"[{completed}/{total_tasks}] [异常] 线程崩溃: {e}")
                    sys.stdout.flush()
                fail_count += 1
                continue

            with print_lock:
                print(f"[{completed}/{total_tasks}] {msg}")
                sys.stdout.flush()

            if success and results:
                all_results.extend(results)
                success_count += 1
            else:
                fail_count += 1

    # 写入Excel（宽表格式：每行一个省份-作物，列为各周VHI，最后一列加折线图）
    if all_results:
        # 国家三字母代码映射（写表用）
        country_code_map = {
            "中国": "CHN", "美国": "USA", "巴西": "BRA", "阿根廷": "ARG",
            "俄罗斯": "RUS", "加拿大": "CAN", "印度": "IND",
            "澳洲": "AUS", "乌克兰": "UKR", "尼日利亚": "NGA",
            "印度尼西亚": "IDN", "马来西亚": "MYS", "哥伦比亚": "COL", "越南": "VNM",
            "法国": "FRA", "德国": "DEU", "波兰": "POL", "罗马尼亚": "ROU",
            "西班牙": "ESP", "意大利": "ITA", "保加利亚": "BGR", "匈牙利": "HUN",
            "丹麦": "DNK", "捷克": "CZE", "英国": "GBR", "奥地利": "AUT", "立陶宛": "LTU",
            "欧盟": "",
            # 省份列是英文，补充英文→代码映射
            "France": "FRA", "Germany": "DEU", "Poland": "POL", "Romania": "ROU",
            "Spain": "ESP", "Italy": "ITA", "Bulgaria": "BGR", "Hungary": "HUN",
            "Denmark": "DNK", "Czech": "CZE", "United Kingdom": "GBR", "Austria": "AUT", "Lithuania": "LTU",
        }
        update_date = datetime.now().strftime('%Y-%m-%d')

        # 收集所有周次并转为宽表（只保留最新的6周数据）
        pivot = defaultdict(dict)
        all_weeks = set()
        for r in all_results:
            key = (r["国家"], r["省份"], r["作物"])
            w = r["周"]
            all_weeks.add(w)
            pivot[key][w] = round(r["VHI"], 2)

        all_weeks = sorted(all_weeks)
        # 只保留最新的6周
        if len(all_weeks) > 6:
            all_weeks = all_weeks[-6:]
            # 过滤pivot中只保留这6周的数据
            for key in list(pivot.keys()):
                pivot[key] = {w: pivot[key].get(w, "") for w in all_weeks}
        
        week_labels = [f"W{w}" for w in all_weeks]
        fieldnames = ["国家", "代码", "省份", "作物"] + week_labels + ["更新日期", "周度走势"]

        # 生成Excel文件
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.marker import Marker
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "VHI_最近6周汇总"
        
        # 写入表头
        for col_idx, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # 设置表头样式：深蓝色背景，白色字体，加粗
            cell.fill = openpyxl.styles.PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        row_idx = 2
        for key in sorted(pivot.keys(), key=lambda k: (k[2], k[0], k[1])):
            country, province, crop = key
            ws.cell(row=row_idx, column=1, value=country)
            # EU行：国家代码从省份名（实际国名）查
            ws.cell(row=row_idx, column=2, value=country_code_map.get(province if country == "欧盟" else country, ""))
            ws.cell(row=row_idx, column=3, value=province)
            ws.cell(row=row_idx, column=4, value=crop)
            
            # 写入周数据
            for col_idx, w in enumerate(all_weeks, 5):
                value = pivot[key].get(w, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # 根据数值设置单元格颜色
                if value != "":
                    if value < 20:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif 20 <= value <= 40:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value > 80:
                        cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            ws.cell(row=row_idx, column=len(all_weeks)+5, value=update_date)
            row_idx += 1
        
        # 设置列宽自适应
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 为数据区域添加边框
        from openpyxl.styles import Border, Side
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=len(fieldnames)):
            for cell in row:
                cell.border = thin_border
        
        # 设置交替行色（仅非数据列，避免覆盖VHI色标）
        data_col_start = 5
        data_col_end = 4 + len(all_weeks)
        for row in range(2, row_idx):
            if row % 2 == 0:
                for col in range(1, len(fieldnames)+1):
                    # 跳过数据列(已由VHI色标着色)和图表列
                    if data_col_start <= col <= data_col_end:
                        continue
                    ws.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 在最后一列添加迷你折线图（无坐标轴，红色折线+圆点标记）
        for row in range(2, row_idx):
            # 创建迷你折线图
            chart = LineChart()
            chart.width = 8
            chart.height = 2
            chart.title = None
            chart.legend = None
            chart.style = 2
            
            # 添加数据系列
            values = Reference(ws, min_col=5, min_row=row, max_col=4+len(all_weeks), max_row=row)
            chart.add_data(values, from_rows=True, titles_from_data=False)
            
            # 设置折线样式：红色实线，圆形标记
            s = chart.series[0]
            s.graphicalProperties.line.solidFill = "E00000"
            s.graphicalProperties.line.width = 20000
            
            marker = Marker(symbol="circle", size=5)
            marker.graphicalProperties.solidFill = "E00000"
            marker.graphicalProperties.line.solidFill = "E00000"
            s.marker = marker
            
            # 隐藏网格线和坐标轴标签
            chart.y_axis.majorGridlines = None
            chart.y_axis.delete = True
            chart.x_axis.majorGridlines = None
            chart.x_axis.delete = True
            
            # 将图表添加到最后一列
            chart_cell = ws.cell(row=row, column=len(fieldnames))
            ws.add_chart(chart, chart_cell.coordinate)
        
        # 保存Excel文件
        wb.save(summary_path)

        print(f"\n汇总Excel文件: {summary_path}")
        print(f"共 {len(pivot)} 行, {len(all_weeks)} 周列")

    print(f"\n{'='*60}")
    print(f"完成! 成功: {success_count}, 失败: {fail_count}, 跳过: {len(skip_records)}")
    print(f"{'='*60}")

    return summary_path if all_results else None


if __name__ == "__main__":
    main()
